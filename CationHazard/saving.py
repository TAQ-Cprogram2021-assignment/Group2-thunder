# 从openpyxl库导入load_workbook函数
from openpyxl import load_workbook


class Saving:
    # 输入方法
    def __init__(self):
        pass

    # 用来存入数据的函数,其中value0为需要传入的变量，b为需要导入的单元格。
    def data_input(self, value0, b):
        # 打开”记录“的工作簿，获取活动工作表
        performance_wb = load_workbook('materials/cation hazard.xlsx')
        performance_ws = performance_wb.active
        # 单元格的输入
        performance_ws[b].value = value0
        # 保存并且关闭excel
        performance_wb.save('materials/cation hazard.xlsx')

    # 用来导出数据的函数,其中self为需要导出所存储的变量，b为需要导出的单元格。
    def data_output(self, b):
        # 打开”记录“的工作簿，获取活动工作表
        performance_wb = load_workbook('materials/cation hazard.xlsx')
        performance_ws = performance_wb.active
        # 单元格的输出
        a = performance_ws[b].value
        # 关闭excel
        performance_wb.save('materials/cation hazard.xlsx')
        return a

    # 基础类
    # 最高得分记录输入
    def highest_score_input(self, val):
        self.data_input(val, 'B2')
        return self

    # 最高得分记录的输出
    def highest_score_output(self):
        a = self.data_output('B2')
        return a

    # 金币输入
    def golden_coin_input(self, val):
        self.data_input(val, 'C2')
        return self

    # 金币的输出
    def golden_coin_output(self):
        a = self.data_output('C2')
        return a

    # 等级的输入
    def level_input(self, val):
        self.data_input(val, 'E2')
        return self

    # 等级的输出
    def level_output(self):
        a = self.data_output('E2')
        return a

    # 天赋类（被动）
    # 子弹等级的输出
    def bullet_level_output(self):
        a = self.data_output('B5')
        return a

    # 子弹等级的输入
    def bullet_level_input(self, val):
        self.data_input(val, 'B5')
        return self

    # 穿透性（概率）的输入
    def blood_level_input(self, val):
        self.data_input(val, 'D5')
        return self

    # 穿透性（概率）的输出
    def blood_level_output(self):
        a = self.data_output('D5')
        return a

    # 下面是ssk让我补加的:经验值的输入输出
    def exp_input(self, val):
        self.data_input(val, 'F2')
        return self

    def exp_output(self):
        a = self.data_output('F2')
        return a
